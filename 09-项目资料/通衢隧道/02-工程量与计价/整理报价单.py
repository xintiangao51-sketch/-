import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

# 读取原始文件
file_path = r'C:\Users\高\Desktop\通衢隧道洞口坡面治理工程\通衢隧道_劳务清包工报价单_不含主材_20260529.xlsx'
output_path = r'C:\Users\高\Desktop\通衢隧道_劳务清包工报价单_整理版_20260530.xlsx'

# 加载工作簿
wb = load_workbook(file_path)

# 遍历所有工作表，整理格式
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # 调整列宽，删除空列
    max_col = sheet.max_column
    cols_to_delete = []
    for col in range(1, max_col + 1):
        is_empty = True
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip() != '':
                is_empty = False
                break
        if is_empty:
            cols_to_delete.append(col)
    # 从后往前删，避免索引错误
    for col in reversed(cols_to_delete):
        sheet.delete_cols(col)
    
    # 设置字体和对齐
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(name='微软雅黑', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            # 数值格式设置
            if isinstance(cell.value, (int, float)):
                if cell.value > 1000:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                else:
                    cell.number_format = numbers.FORMAT_NUMBER_00
    
    # 调整列宽
    for col_idx in range(1, sheet.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_idx)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[col_letter].width = adjusted_width

# 保存整理后的文件
wb.save(output_path)
print(f"整理后的报价单已保存到：{output_path}")
