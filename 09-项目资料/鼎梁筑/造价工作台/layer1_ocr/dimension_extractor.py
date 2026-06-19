#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Layer 1 — 尺寸标注提取
功能: 从OCR文本中提取尺寸标注并计算工程量
"""

import re
import logging
from typing import List, Optional
from dataclasses import dataclass

logger = logging.getLogger(__name__)

try:
    from shared.utils import extract_dimensions, classify_component, normalize_unit
except ImportError:
    import sys, os
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
    from shared.utils import extract_dimensions, classify_component, normalize_unit

from shared.models import DimensionResult, CategoryType


class DimensionExtractor:
    """
    尺寸标注提取器
    从OCR识别的文本中提取:
    - 截面尺寸 (宽×高)
    - 长度标注
    - 混凝土标号
    - 钢筋规格
    - 自动计算工程量
    """

    def __init__(self):
        self.results: List[DimensionResult] = []

    def extract_from_text(self, text: str, page: int = 0) -> List[DimensionResult]:
        """
        从单页文本提取所有尺寸标注

        :param text: OCR文本
        :param page: 页码
        :return: DimensionResult列表
        """
        results = []
        lines = text.split("\n")

        for line in lines:
            line = line.strip()
            if len(line) < 5:
                continue

            dims = extract_dimensions(line)
            if not dims:
                continue

            comp_type = classify_component(line)
            try:
                category = CategoryType(comp_type)
            except ValueError:
                category = CategoryType.OTHER

            # 构造成果
            dim = self._build_dimension(line, dims, category, page)
            if dim:
                results.append(dim)
                self.results.append(dim)

        return results

    def _build_dimension(self, text: str, dims: list, category: CategoryType,
                        page: int) -> Optional[DimensionResult]:
        """从提取的尺寸构建DimensionResult"""
        result = DimensionResult(
            text=text[:100],
            category=category,
            page=page,
            confidence=0.0,
        )

        for d in dims:
            dim_type = d["type"]
            if dim_type == "section" and d["groups"]:
                if len(d["groups"]) >= 2:
                    result.width = float(d["groups"][0])
                    result.height = float(d["groups"][1])
            elif dim_type == "length" and d["groups"]:
                result.length = float(d["groups"][0])
            elif dim_type == "concrete_grade" and d["groups"]:
                result.concrete_grade = d["groups"][0]
            elif dim_type == "rebar" and d["groups"]:
                result.rebar_spec = d["value"]
            elif dim_type == "pipe_diameter" and d["groups"]:
                result.rebar_spec = f"DN{d['groups'][0]}"

        # 自动计算工程量
        self._calc_quantity(result)

        return result

    def _calc_quantity(self, dim: DimensionResult):
        """根据尺寸自动计算工程量"""
        # 梁: 宽×高 → 截面积
        if dim.category == CategoryType.BEAM and dim.width and dim.height:
            dim.unit = "m³/m"  # 每米体积
            dim.quantity = (dim.width * dim.height) / 1_000_000  # mm² → m²

        # 柱: 宽×高×长 → 体积
        elif dim.category == CategoryType.COLUMN and dim.width and dim.height:
            dim.unit = "m³"
            h = dim.length if dim.length else 3000  # 默认层高3m
            dim.quantity = (dim.width * dim.height * h) / 1_000_000_000

        # 板: 面积 → 默认厚度100mm
        elif dim.category == CategoryType.SLAB:
            dim.unit = "m³"
            thick = dim.height if dim.height else 100  # 默认板厚
            if dim.width and dim.length:
                dim.quantity = (dim.width * dim.length * thick) / 1_000_000_000

        # 基础: 体积
        elif dim.category == CategoryType.FOUNDATION:
            dim.unit = "m³"
            if dim.width and dim.length:
                h = dim.height if dim.height else 500
                dim.quantity = (dim.width * dim.length * h) / 1_000_000_000

        # 钢筋: 按重量 (7850kg/m³)
        elif dim.category == CategoryType.REBAR:
            dim.unit = "kg/m"
            if dim.rebar_spec:
                dia_match = re.search(r"Φ?\s*(\d+)", dim.rebar_spec)
                if dia_match:
                    dia_mm = float(dia_match.group(1))
                    dim.quantity = dia_mm ** 2 * 0.00617  # kg/m 公式

        # 墙: 面积
        elif dim.category == CategoryType.WALL:
            dim.unit = "m³"
            thick = dim.width if dim.width else 200
            if dim.length:
                h = dim.height if dim.height else 3000
                dim.quantity = (thick * dim.length * h) / 1_000_000_000

    def extract_from_pages(self, pages: List[dict]) -> List[DimensionResult]:
        """从多页OCR结果提取"""
        all_dims = []
        for page_data in pages:
            dims = self.extract_from_text(
                page_data["text"],
                page_data.get("page", 0)
            )
            all_dims.extend(dims)
        return all_dims

    def group_by_category(self) -> dict:
        """按构件类型分组"""
        from collections import defaultdict
        groups = defaultdict(list)
        for dim in self.results:
            groups[dim.category.value].append(dim)
        return dict(groups)

    def to_dataframe(self):
        """转换为DataFrame (需要pandas)"""
        try:
            import pandas as pd
            return pd.DataFrame([
                {
                    "构件类型": d.category.value,
                    "宽(mm)": d.width,
                    "高(mm)": d.height,
                    "长(mm)": d.length,
                    "混凝土": d.concrete_grade,
                    "钢筋": d.rebar_spec,
                    "单位": d.unit,
                    "工程量": round(d.quantity, 4),
                    "页码": d.page,
                    "来源文本": d.text[:50],
                }
                for d in self.results
            ])
        except ImportError:
            return None

    def export_excel(self, filepath: str):
        """导出到Excel"""
        from shared.excel_io import write_boq_sheet

        headers = ["构件类型", "宽(mm)", "高(mm)", "长(mm)", "混凝土标号",
                    "钢筋规格", "单位", "工程量", "页码", "来源文本"]
        rows = []
        for d in self.results:
            rows.append([
                d.category.value, d.width, d.height, d.length,
                d.concrete_grade, d.rebar_spec, d.unit,
                round(d.quantity, 4), d.page, d.text[:60],
            ])

        # 使用excel_io写入
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "图纸工程量提取"

        # 标题
        ws.merge_cells('A1:J1')
        ws['A1'].value = "图纸工程量提取结果"
        ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)

        # 表头
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # 数据
        for row_idx, row_data in enumerate(rows, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='微软雅黑', size=10)
                cell.border = thin_border

        wb.save(filepath)
        logger.info(f"工程量提取结果已保存: {filepath}")
        return filepath
