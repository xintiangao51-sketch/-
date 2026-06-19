#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
鼎梁筑造价工作台 — Excel读写模块
支持: openpyxl 读写、GB50500-2013格式清单、报价表模板
"""

import os
import logging
from typing import List, Optional, Dict, Any
from pathlib import Path

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl 未安装，Excel功能不可用")


# ============ 样式常量 ============

HEADER_FONT = Font(name='微软雅黑', size=11, bold=True)
DATA_FONT = Font(name='微软雅黑', size=10)
TITLE_FONT = Font(name='微软雅黑', size=14, bold=True)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT_WHITE = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 偏差颜色
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# BOQ列定义
BOQ_HEADERS = [
    "清单编码", "项目名称", "项目特征", "单位", "工程量",
    "综合单价", "合价", "人工费", "材料费", "机械费",
    "管理费", "利润", "匹配定额", "置信度"
]


def ensure_openpyxl():
    """确保openpyxl可用"""
    if not HAS_OPENPYXL:
        raise ImportError("请安装 openpyxl: pip install openpyxl --break-system-packages")


def read_excel(filepath: str, sheet_name: str = None, header_row: int = 1) -> List[Dict]:
    """
    读取Excel为字典列表
    :param filepath: Excel文件路径
    :param sheet_name: 工作表名(None=第一个)
    :param header_row: 表头行号(1-based)
    :return: [{col_name: value}, ...]
    """
    ensure_openpyxl()
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 读表头
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        headers.append(str(cell_value).strip() if cell_value else f"col_{col}")

    # 读数据
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        has_data = False
        for col, header in enumerate(headers, 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                has_data = True
            row_data[header] = value
        if has_data:
            rows.append(row_data)

    wb.close()
    return rows


def write_boq_sheet(filepath: str, sheet_data, sheet_name: str = "工程量清单"):
    """
    写入BOQ清单到Excel (GB50500-2013格式)
    :param filepath: 输出路径
    :param sheet_data: BOQSheet对象 或 items列表
    :param sheet_name: 工作表名
    """
    ensure_openpyxl()
    from shared.models import BOQSheet, BOQItem

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 标题行
    if isinstance(sheet_data, BOQSheet):
        project_name = sheet_data.project_name
        items = sheet_data.items
    else:
        project_name = ""
        items = sheet_data

    ws.merge_cells('A1:N1')
    cell = ws['A1']
    cell.value = project_name or "工程量清单"
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN

    # 表头
    for col_idx, header in enumerate(BOQ_HEADERS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 数据行
    for row_idx, item in enumerate(items, 4):
        row_data = item.to_row() if isinstance(item, BOQItem) else [
            item.get("code", ""), item.get("name", ""), item.get("feature", ""),
            item.get("unit", ""), item.get("quantity", 0),
            item.get("unit_price", 0), item.get("total_price", 0),
            item.get("labor_cost", 0), item.get("material_cost", 0),
            item.get("machinery_cost", 0), item.get("management_fee", 0),
            item.get("profit", 0), item.get("matched_quota", ""),
            item.get("match_confidence", 0),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN if col_idx <= 3 else CENTER_ALIGN

    # 列宽
    col_widths = [16, 28, 30, 8, 12, 12, 14, 12, 12, 12, 10, 10, 14, 8]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 汇总行
    if isinstance(sheet_data, BOQSheet):
        last_row = len(items) + 4
        sheet_data.recalc()
        summary = [
            ("合计", sheet_data.total_amount),
            ("人工费合计", sheet_data.total_labor),
            ("材料费合计", sheet_data.total_material),
            ("机械费合计", sheet_data.total_machinery),
            ("管理费合计", sheet_data.total_management),
            ("利润合计", sheet_data.total_profit),
        ]
        for i, (label, value) in enumerate(summary):
            row = last_row + 1 + i
            ws.cell(row=row, column=1, value=label).font = Font(name='微软雅黑', size=10, bold=True)
            ws.cell(row=row, column=7, value=value).font = Font(name='微软雅黑', size=10, bold=True)

    wb.save(filepath)
    logger.info(f"BOQ清单已保存: {filepath}")
    return filepath


def write_diff_report(filepath: str, diff_records: List, project_name: str = "") -> str:
    """
    写入差异核验报告
    :param filepath: 输出路径
    :param diff_records: DiffRecord列表
    """
    ensure_openpyxl()
    from shared.models import DiffRecord

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "差异核验"

    # 标题
    ws.merge_cells('A1:G1')
    ws['A1'].value = f"{project_name} - 图纸vs清单差异核验报告"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER_ALIGN

    # 表头
    headers = ["清单编码", "项目名称", "图纸量", "清单量", "偏差率(%)", "偏差类型", "风险等级"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 数据
    for row_idx, dr in enumerate(diff_records, 4):
        if isinstance(dr, DiffRecord):
            row_data = [dr.code, dr.name, dr.drawing_qty, dr.boq_qty,
                       dr.deviation, dr.deviation_type, dr.risk_level.value]
        else:
            row_data = [dr.get(k, "") for k in ["code", "name", "drawing_qty", "boq_qty",
                         "deviation", "deviation_type", "risk_level"]]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

            # 偏差率着色
            if col_idx == 5 and isinstance(value, (int, float)):
                abs_val = abs(value)
                if abs_val <= 5:
                    cell.fill = GREEN_FILL
                elif abs_val <= 15:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = RED_FILL

    # 列宽
    for col_idx, w in enumerate([16, 28, 12, 12, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(filepath)
    logger.info(f"差异报告已保存: {filepath}")
    return filepath


def read_boq_from_excel(filepath: str, sheet_name: str = None) -> List:
    """
    从Excel读取BOQ清单并转换为BOQItem列表
    """
    from shared.models import BOQItem

    rows = read_excel(filepath, sheet_name)
    items = []
    for row in rows:
        try:
            item = BOQItem(
                code=str(row.get("清单编码", row.get("编码", ""))).strip(),
                name=str(row.get("项目名称", row.get("名称", ""))).strip(),
                feature=str(row.get("项目特征", row.get("特征", ""))).strip(),
                unit=str(row.get("单位", "")).strip(),
                quantity=float(row.get("工程量", 0) or 0),
                unit_price=float(row.get("综合单价", 0) or 0),
                total_price=float(row.get("合价", 0) or 0),
            )
            items.append(item)
        except (ValueError, TypeError):
            continue
    return items


def read_ocr_results(filepath: str) -> List:
    """
    读取OCR提取的工程量Excel
    返回 DimensionResult 列表
    """
    from shared.models import DimensionResult, CategoryType

    rows = read_excel(filepath)
    results = []
    for row in rows:
        category_str = str(row.get("构件类型", row.get("类别", "其他"))).strip()
        try:
            category = CategoryType(category_str)
        except ValueError:
            category = CategoryType.OTHER

        dim = DimensionResult(
            text=str(row.get("原始文本", "")).strip(),
            width=float(row.get("宽", 0) or 0),
            height=float(row.get("高", 0) or 0),
            length=float(row.get("长", 0) or 0),
            concrete_grade=str(row.get("混凝土标号", "")).strip(),
            rebar_spec=str(row.get("钢筋规格", "")).strip(),
            category=category,
            unit=str(row.get("单位", "")).strip(),
            quantity=float(row.get("工程量", 0) or 0),
            confidence=float(row.get("置信度", 0) or 0),
        )
        results.append(dim)
    return results
