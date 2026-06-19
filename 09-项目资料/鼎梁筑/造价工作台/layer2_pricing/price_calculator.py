#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Layer 2 — 综合单价计算器
功能: 定额套价 → 人材机拆分 → 取费 → 综合单价 + 合价
"""

import os
import logging
from typing import List, Optional

logger = logging.getLogger(__name__)

try:
    from shared.models import (
        BOQItem, BOQSheet, QuotaItem,
        FeeConfig, Region, create_fee_config
    )
    from shared.utils import get_labor_rate, get_fee_rate
except ImportError:
    import sys
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
    from shared.models import (
        BOQItem, BOQSheet, QuotaItem,
        FeeConfig, Region, create_fee_config
    )
    from shared.utils import get_labor_rate, get_fee_rate


class PriceCalculator:
    """
    综合单价计算器

    计算流程:
    1. 读入工程量清单
    2. 匹配定额子目
    3. 套取定额人材机消耗量 × 工程量
    4. 按费率计算: 管理28% + 利润18% + 规费22% + 增值税9%
    5. 输出综合单价 + 合价
    """

    def __init__(self, region: str = "乌鲁木齐", region_category: str = "一类"):
        """
        :param region: 地区(乌鲁木齐/伊犁/喀什/石河子/阿克苏/昌吉)
        :param region_category: 人工类别(一类/二类/三类)
        """
        self.region = region
        self.region_category = region_category
        self.labor_rate = get_labor_rate(region, region_category) or 130.0
        self.fee_config = create_fee_config()

        # 喀什高海拔调整
        if region == "喀什":
            self.fee_config.management_rate = 0.30  # 高海拔+2%

        logger.info(f"计价参数: 地区={region}, 人工={self.labor_rate}元/工日, "
                    f"管理={self.fee_config.management_rate*100:.0f}%, "
                    f"利润={self.fee_config.profit_rate*100:.0f}%, "
                    f"规费={self.fee_config.regulation_rate*100:.0f}%, "
                    f"税={self.fee_config.vat_rate*100:.0f}%")

    def calc_from_quota(self, boq_item: BOQItem, quota_item: QuotaItem,
                        labor_override: float = None) -> BOQItem:
        """
        根据定额子目计算单个清单项的综合单价

        :param boq_item: BOQ清单项(含工程量和匹配定额)
        :param quota_item: 匹配的定额子目
        :param labor_override: 覆盖人工单价(None=使用默认)
        :return: 更新后的BOQItem(含完整费用拆解)
        """
        labor_rate = labor_override or self.labor_rate
        quantity = boq_item.quantity

        # --- 步骤1: 套定额消耗量 ---
        # 定额基价按比例折算到清单工程量
        # 折算系数 = 清单工程量 / 定额单位量
        qty_ratio = self._calc_qty_ratio(boq_item.unit, quota_item.unit, quantity)

        # 人工费 (人工消耗量 × 人工单价 × 折算系数)
        boq_item.labor_cost = quota_item.labor_qty * labor_rate * qty_ratio

        # 材料费 (主材 + 辅材)
        boq_item.material_cost = (quota_item.main_material_cost +
                                  quota_item.aux_material_cost) * qty_ratio

        # 机械费
        boq_item.machinery_cost = quota_item.machinery_cost * qty_ratio

        # --- 步骤2: 计取管理费和利润 ---
        boq_item.management_fee = self.fee_config.calc_management(
            boq_item.labor_cost, boq_item.machinery_cost
        )
        boq_item.profit = self.fee_config.calc_profit(
            boq_item.labor_cost, boq_item.machinery_cost
        )

        # --- 步骤3: 综合单价 ---
        # 综合单价 = 人工 + 材料 + 机械 + 管理 + 利润
        boq_item.unit_price = (
            boq_item.labor_cost + boq_item.material_cost +
            boq_item.machinery_cost + boq_item.management_fee +
            boq_item.profit
        )

        # --- 步骤4: 合价 ---
        boq_item.total_price = boq_item.unit_price * boq_item.quantity

        return boq_item

    def _calc_qty_ratio(self, boq_unit: str, quota_unit: str,
                        boq_qty: float) -> float:
        """
        计算工程量折算系数

        定额单位通常是 10m³/100m²/100m/t 等放大单位
        boq清单量是实际工程量

        :return: 折算系数
        """
        # 解析定额单位的放大倍数
        quota_multiplier = 1.0
        quota_base = quota_unit

        # 常见放大单位: 10m³, 100m², 100m, t
        import re
        mult_match = re.match(r"(\d+)\s*(.+)", quota_unit)
        if mult_match:
            quota_multiplier = float(mult_match.group(1))
            quota_base = mult_match.group(2)

        # 单位相同 → 直接使用工程量 ÷ 放大倍数
        # 单位不同 → 需要近似转换

        # 简化: 假设单位基本可换算
        # 实际项目中需要更精细的单位转换表
        unit_normalized_boq = boq_unit.replace("³", "3").replace("²", "2")
        unit_normalized_quota = quota_base.replace("³", "3").replace("²", "2")

        if unit_normalized_boq == unit_normalized_quota:
            return boq_qty / quota_multiplier
        elif "m3" in unit_normalized_boq and "m3" in unit_normalized_quota:
            return boq_qty / quota_multiplier
        elif "m2" in unit_normalized_boq and "m2" in unit_normalized_quota:
            return boq_qty / quota_multiplier
        elif "m" in unit_normalized_boq and "m" in unit_normalized_quota:
            return boq_qty / quota_multiplier
        elif "t" in unit_normalized_boq and "t" in unit_normalized_quota:
            return boq_qty / quota_multiplier
        else:
            # 单位不匹配，保守处理
            logger.warning(f"单位不匹配: BOQ={boq_unit}, Quota={quota_unit}, 使用1:1折算")
            return boq_qty / quota_multiplier

    def calc_sheet(self, sheet: BOQSheet, quota_db) -> BOQSheet:
        """
        计算整张清单表

        :param sheet: 工程量清单表
        :param quota_db: QuotaDatabase实例
        :return: 更新后的BOQSheet(含完整取费)
        """
        matched_count = 0
        total = len(sheet.items)

        for i, item in enumerate(sheet.items):
            if item.matched_quota:
                quota = quota_db.get(item.matched_quota)
                if quota:
                    self.calc_from_quota(item, quota)
                    matched_count += 1
                else:
                    logger.warning(f"定额不存在: {item.matched_quota} for {item.code}")
            else:
                logger.debug(f"未匹配定额: {item.code} {item.name}")

            from shared.utils import progress_bar
            progress_bar(i + 1, total, "组价计算")

        sheet.recalc()
        logger.info(f"组价完成: {matched_count}/{total} 项 ({(matched_count/total*100):.1f}%)")
        return sheet

    def calc_total_with_fees(self, sheet: BOQSheet) -> dict:
        """
        计算含措施费+规费+税金的完整造价

        :return: {分部分项合计, 措施费, 安全文明, 规费, 税前合计, 增值税, 含税总造价}
        """
        sheet.recalc()
        labor_total = sheet.total_labor

        # 措施费 (≈分部分项×4.5%)
        measure_fee = sheet.total_amount * self.fee_config.measure_rate

        # 安全文明施工费 (≈分部分项×2.5%)
        safety_fee = sheet.total_amount * self.fee_config.safety_rate

        # 规费 (人工费×22%)
        regulation = self.fee_config.calc_regulation(labor_total)

        # 税前合计
        pre_tax = sheet.total_amount + measure_fee + safety_fee + regulation

        # 增值税 (9%)
        vat = self.fee_config.calc_vat(pre_tax)

        # 含税总造价
        total_with_tax = pre_tax + vat

        return {
            "分部分项合计": round(sheet.total_amount, 2),
            "其中人工费": round(labor_total, 2),
            "其中材料费": round(sheet.total_material, 2),
            "其中机械费": round(sheet.total_machinery, 2),
            "其中管理费": round(sheet.total_management, 2),
            "其中利润": round(sheet.total_profit, 2),
            "措施项目费": round(measure_fee, 2),
            "安全文明施工费": round(safety_fee, 2),
            "规费": round(regulation, 2),
            "税前合计": round(pre_tax, 2),
            "增值税(9%)": round(vat, 2),
            "含税总造价": round(total_with_tax, 2),
        }

    def generate_quote_sheet(self, boq_items: List[BOQItem],
                             project_name: str,
                             quota_db) -> BOQSheet:
        """
        一键生成报价表

        :param boq_items: 清单项列表
        :param project_name: 项目名称
        :param quota_db: 定额数据库
        :return: 完整的BOQSheet
        """
        from layer2_pricing.quota_matcher import QuotaMatcher

        matcher = QuotaMatcher()
        # 如果传入了外部db，使用外部db
        if quota_db and quota_db.items:
            matcher.database = quota_db

        sheet = BOQSheet(
            project_name=project_name,
            sheet_name="清单报价",
            items=boq_items,
        )

        # 匹配定额
        for item in sheet.items:
            matcher.match_boq_item(item)

        # 计算单价
        self.calc_sheet(sheet, matcher.database)

        return sheet

    def export_cost_summary(self, filepath: str, sheet: BOQSheet) -> str:
        """导出造价汇总表"""
        from shared.excel_io import write_boq_sheet, ensure_openpyxl
        ensure_openpyxl()
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        result = self.calc_total_with_fees(sheet)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "造价汇总"

        # 标题
        ws.merge_cells('A1:C1')
        ws['A1'].value = f"{sheet.project_name} — 造价汇总"
        ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)

        # 汇总表
        headers = ["费用项", "金额(元)", "备注"]
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = Font(name='微软雅黑', size=10, bold=True)
            cell.border = thin_border

        summary_items = [
            ("分部分项合计", result["分部分项合计"], f"{len(sheet.items)}项"),
            ("  人工费", result["其中人工费"], f"{self.labor_rate}元/工日"),
            ("  材料费", result["其中材料费"], "含主材+辅材"),
            ("  机械费", result["其中机械费"], ""),
            ("  管理费", result["其中管理费"], f"{self.fee_config.management_rate*100:.0f}%×(人+机)"),
            ("  利润", result["其中利润"], f"{self.fee_config.profit_rate*100:.0f}%×(人+机)"),
            ("措施项目费", result["措施项目费"], f"≈{self.fee_config.measure_rate*100:.0f}%"),
            ("安全文明施工费", result["安全文明施工费"], f"≈{self.fee_config.safety_rate*100:.0f}%"),
            ("规费", result["规费"], f"{self.fee_config.regulation_rate*100:.0f}%×人工"),
            ("税前合计", result["税前合计"], ""),
            ("增值税", result["增值税(9%)"], "9%"),
            ("含税总造价", result["含税总造价"], "★"),
        ]

        for row_idx, (label, amount, note) in enumerate(summary_items, 4):
            ws.cell(row=row_idx, column=1, value=label).border = thin_border
            cell = ws.cell(row=row_idx, column=2, value=amount)
            cell.border = thin_border
            cell.number_format = '#,##0.00'
            ws.cell(row=row_idx, column=3, value=note).border = thin_border

            # 加粗关键行
            if label in ("分部分项合计", "含税总造价"):
                ws.cell(row=row_idx, column=1).font = Font(name='微软雅黑', size=11, bold=True)
                ws.cell(row=row_idx, column=2).font = Font(name='微软雅黑', size=11, bold=True)

        # 参数说明
        param_row = len(summary_items) + 5
        ws.cell(row=param_row, column=1, value="计价参数").font = Font(name='微软雅黑', size=11, bold=True)
        params = [
            f"定额: 新疆2024市政消耗量定额",
            f"地区: {self.region}",
            f"人工基价: {self.labor_rate}元/工日 ({self.region_category}类)",
            f"管理费: {self.fee_config.management_rate*100:.0f}% | 利润: {self.fee_config.profit_rate*100:.0f}%",
            f"规费: {self.fee_config.regulation_rate*100:.0f}% | 增值税: {self.fee_config.vat_rate*100:.0f}%",
        ]
        for i, param in enumerate(params):
            ws.cell(row=param_row + 1 + i, column=1, value=param)

        wb.save(filepath)
        logger.info(f"造价汇总已保存: {filepath}")
        return filepath
