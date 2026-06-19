#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从projects.json生成新疆工程造价项目汇总Excel表
功能：
1. 表头对应所有提取字段
2. 自动调整列宽
3. 对缺失的下浮率/单方造价用公式自动计算
4. 添加自动筛选
"""

import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.filters import AutoFilter

def load_projects(json_path):
    """加载JSON数据"""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data

def get_all_fields(projects):
    """获取所有字段，确保表头包含所有提取字段"""
    all_fields = set()
    for project in projects:
        all_fields.update(project.keys())
    # 按照常用顺序排序，把重要字段放前面
    priority_fields = [
        '项目名称', '项目地点', '项目类型', '建筑面积', '结构类型', 
        '总造价', '单方造价', '下浮率', '中标时间', '建设单位', 
        '施工单位', '咨询单位', '备注'
    ]
    # 把优先字段放前面，剩下的放后面
    sorted_fields = []
    for field in priority_fields:
        if field in all_fields:
            sorted_fields.append(field)
            all_fields.remove(field)
    # 添加剩余字段
    sorted_fields.extend(sorted(all_fields))
    return sorted_fields

def calculate_column_width(ws):
    """自动调整列宽"""
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # 中文字符宽度计算
                cell_length = 0
                for char in str(cell.value):
                    if ord(char) > 127:
                        cell_length += 2
                    else:
                        cell_length += 1
                if cell_length > max_length:
                    max_length = cell_length
        # 给一点额外边距
        adjusted_width = max_length + 2
        # 限制最大宽度
        if adjusted_width > 50:
            adjusted_width = 50
        elif adjusted_width < 10:
            adjusted_width = 10
        ws.column_dimensions[column].width = adjusted_width

def create_excel(projects, output_path, fields):
    """创建Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "新疆工程造价项目汇总"
    
    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 写入表头
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # 写入数据
    for row_idx, project in enumerate(projects, 2):
        for col_idx, field in enumerate(fields, 1):
            value = project.get(field, None)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 数字类型右对齐，文本左对齐
            if isinstance(value, (int, float)):
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
            # 处理缺失值和公式计算
            if field == '单方造价' and (value is None or value == ''):
                # 查找总造价和建筑面积的列
                total_price_col = None
                area_col = None
                for i, f in enumerate(fields, 1):
                    if f == '总造价':
                        total_price_col = i
                    if f == '建筑面积':
                        area_col = i
                # 如果都存在，添加公式：总造价/建筑面积
                if total_price_col and area_col:
                    total_price_cell = f"{get_column_letter(total_price_col)}{row_idx}"
                    area_cell = f"{get_column_letter(area_col)}{row_idx}"
                    cell.value = f"=IF({area_cell}=0,0,{total_price_cell}/{area_cell})"
            
            if field == '下浮率' and (value is None or value == ''):
                # 查找控制价和中标价的列（不同命名方式）
                control_price_col = None
                bid_price_col = None
                control_names = ['控制价', '招标控制价', '预算价']
                bid_names = ['中标价', '成交价格', '总造价']
                
                for i, f in enumerate(fields, 1):
                    if f in control_names and control_price_col is None:
                        control_price_col = i
                    if f in bid_names and bid_price_col is None:
                        bid_price_col = i
                
                # 如果都存在，添加公式：(控制价-中标价)/控制价*100%
                if control_price_col and bid_price_col:
                    control_cell = f"{get_column_letter(control_price_col)}{row_idx}"
                    bid_cell = f"{get_column_letter(bid_price_col)}{row_idx}"
                    cell.value = f"=IF({control_cell}=0,0,({control_cell}-{bid_cell})/{control_cell})"
                    cell.number_format = '0.00%'
    
    # 添加自动筛选
    last_col = get_column_letter(len(fields))
    last_row = len(projects) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 自动调整列宽
    calculate_column_width(ws)
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}")
    print(f"共处理 {len(projects)} 个项目，{len(fields)} 个字段")

def main():
    """主函数"""
    json_path = r"D:\知识库\output\projects.json"
    output_path = r"D:\知识库\output\新疆工程造价项目汇总.xlsx"
    
    # 检查文件是否存在
    import os
    if not os.path.exists(json_path):
        print(f"错误：找不到输入文件 {json_path}")
        print("请确保projects.json文件存在于output目录下")
        # 创建一个空的示例文件，方便用户后续使用
        print("\n创建空的模板Excel...")
        # 预设常用字段
        default_fields = [
            '项目名称', '项目地点', '项目类型', '建筑面积(㎡)', 
            '结构类型', '控制价(万元)', '中标价(万元)', '总造价(万元)', 
            '单方造价(元/㎡)', '下浮率', '中标时间', '建设单位', 
            '施工单位', '咨询单位', '备注'
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "新疆工程造价项目汇总"
        
        # 写入表头
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx, field in enumerate(default_fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        calculate_column_width(ws)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(default_fields))}1"
        ws.freeze_panes = "A2"
        wb.save(output_path)
        print(f"空模板已创建到: {output_path}")
        print("使用的默认字段:", default_fields)
        return
    
    # 加载数据
    projects = load_projects(json_path)
    print(f"加载了 {len(projects)} 个项目")
    
    # 获取所有字段
    fields = get_all_fields(projects)
    print(f"字段列表: {fields}")
    
    # 创建Excel
    create_excel(projects, output_path, fields)
    print("处理完成！")

if __name__ == '__main__':
    main()
