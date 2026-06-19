import os, re, json, math, csv, statistics
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook, Workbook

root = Path(r'D:/知识库/09-项目资料/哈密国源综合服务中心')
if not root.exists() or not any(root.rglob('*.xlsx')):
    # 回退到桌面路径
    root = Path(r'C:/Users/20752/Desktop/哈密国源综合服务中心项目/5-7国源电力哈密煤电公司综合能源集控服务中心工程施工公开招标/招标文件/附件')
if not root.exists():
    print(f"❌ 路径不存在: {root}")
    print("请修改脚本中的 root 变量为实际路径")
    raise SystemExit(1)
print(f"✅ 扫描路径: {root}")
out_root = Path(r'D:/知识库/09-项目资料/哈密国源综合服务中心/原始附件扫描')
out_root.mkdir(parents=True, exist_ok=True)

money_labels = ['分部分项','措施','单价措施','总价措施','其他项目','暂列','暂估','规费','税金','安全文明','合计','工程造价','单位工程']
code_re = re.compile(r'^(?:\d{9,12}|[A-Z]?\d{6,12}|\d{2}[A-Z]\d{6,})$')

xlsx_files = sorted([p for p in root.rglob('*.xlsx') if not p.name.startswith('~$')])
summary_rows=[]
sheet_rows=[]
bill_rows=[]
errors=[]

def clean(v):
    if v is None: return ''
    if isinstance(v,float) and math.isnan(v): return ''
    return str(v).strip().replace('\n',' ')

def as_num(v):
    if isinstance(v,(int,float)) and not isinstance(v,bool): return float(v)
    if isinstance(v,str):
        s=v.replace(',','').replace('￥','').strip()
        try: return float(s)
        except: return None
    return None

def row_values(ws, r, max_col=None):
    max_col = max_col or min(ws.max_column, 20)
    return [ws.cell(r,c).value for c in range(1, max_col+1)]

def joined(vals):
    return ' | '.join(clean(v) for v in vals if clean(v))

for idx,p in enumerate(xlsx_files,1):
    rel=str(p.relative_to(root)).replace('\\','/')
    rec={'rel':rel,'path':str(p),'status':'ok','sheet_count':0,'sheets':'','project':'','unit':'','file_size':p.stat().st_size}
    try:
        wb = load_workbook(p, data_only=True, read_only=True)
        rec['sheet_count']=len(wb.sheetnames)
        rec['sheets']='; '.join(wb.sheetnames)
        # classify by path
        rec['section']='新能源' if '新能源' in rel else ('明珠' if '明珠' in rel else ('清单汇总' if rel.startswith('清单/') or '/清单/' in rel else ('图纸附件' if '招标图' in rel else '其他')))
        rec['discipline']=''
        parts=rel.split('/')
        for key in ['建筑工程','装饰工程','电气工程','通风空调工程','消防报警工程','消防水工程','弱电及监控工程','给排水工程','太阳能工程','人防工程','道路工程','停车场工程','园林绿化工程','围墙工程','强弱电工程','热力工程','室外工程','集控中心','调度中心']:
            if key in rel:
                rec['discipline']=key
        wb_bill_count=0
        wb_money=[]
        for sname in wb.sheetnames:
            ws=wb[sname]
            srec={'file':rel,'sheet':sname,'max_row':ws.max_row,'max_col':ws.max_column,'bill_items':0,'amounts':''}
            # read top rows for project/unit and amount labels
            maxr=min(ws.max_row or 0, 80)
            maxc=min(ws.max_column or 0, 15)
            amount_hits=[]
            for r in range(1,maxr+1):
                vals=row_values(ws,r,maxc)
                txt=joined(vals)
                if txt and not rec['project'] and '工程名称' in txt:
                    rec['project']=txt[:250]
                if txt and not rec['unit'] and ('单位工程' in txt or '单项工程' in txt):
                    rec['unit']=txt[:250]
                # summary amount rows
                if txt and any(label in txt for label in money_labels):
                    nums=[as_num(v) for v in vals]
                    nums=[n for n in nums if n is not None and abs(n)>0.0001]
                    # ignore pure serial rows, keep largest amount if sensible
                    if nums:
                        amount_hits.append({'r':r,'text':txt[:220],'nums':nums[-5:]})
                # bill item detection: code in first 3 columns + unit/qty later
                cells=[clean(v) for v in vals]
                code=None
                for c in range(min(4,len(cells))):
                    s=cells[c].replace(' ','')
                    if code_re.match(s) and len(s)>=9:
                        code=s; break
                if code:
                    name=''
                    unit=''
                    qty=None
                    # find name as next nonempty text after code
                    found=False
                    for val in cells:
                        if not found and val.replace(' ','')==code:
                            found=True; continue
                        if found and val and not as_num(val) and len(val)>1:
                            name=val; break
                    for v in vals:
                        n=as_num(v)
                        if n is not None and abs(n)>0.000001:
                            # likely quantity if not code and not serial too big maybe just keep first after unit
                            qty=n
                    bill_rows.append({'file':rel,'sheet':sname,'row':r,'code':code,'name':name[:120],'qty_lastnum':qty})
                    srec['bill_items']+=1
            wb_bill_count += srec['bill_items']
            if amount_hits:
                # concise hits
                srec['amounts']=' || '.join([f"R{h['r']}:{h['text']} => {h['nums']}" for h in amount_hits[:12]])
                wb_money.extend(amount_hits)
            sheet_rows.append(srec)
        rec['bill_items']=wb_bill_count
        # file total: choose last/largest amount labelled 合计/工程造价 in amount hits
        candidates=[]
        for h in wb_money:
            txt=h['text']
            if any(k in txt for k in ['合计','工程造价','单位工程','招标控制价','投标报价']):
                for n in h['nums']:
                    if abs(n)>1000: candidates.append((abs(n),n,txt,h['r']))
        if candidates:
            candidates.sort(reverse=True)
            rec['major_amount']=candidates[0][1]
            rec['major_amount_label']=f"R{candidates[0][3]} {candidates[0][2][:180]}"
        else:
            rec['major_amount']=''
            rec['major_amount_label']=''
        summary_rows.append(rec)
    except Exception as e:
        rec['status']='error'; rec['error']=repr(e); rec['bill_items']=0; rec['major_amount']=''; rec['major_amount_label']=''
        summary_rows.append(rec); errors.append({'file':rel,'error':repr(e)})

# write JSON/CSV
(out_root/'xlsx_parse_summary.json').write_text(json.dumps({'files':summary_rows,'sheets':sheet_rows,'bill_rows_sample':bill_rows[:5000],'errors':errors},ensure_ascii=False,indent=2),encoding='utf-8')
for name,rows in [('xlsx_files_summary.csv',summary_rows),('xlsx_sheets_summary.csv',sheet_rows),('xlsx_bill_items_sample.csv',bill_rows[:20000])]:
    if not rows: continue
    keys=list(rows[0].keys())
    with (out_root/name).open('w',encoding='utf-8-sig',newline='') as f:
        w=csv.DictWriter(f,fieldnames=keys,extrasaction='ignore'); w.writeheader(); w.writerows(rows)

# write workbook report
wbout=Workbook()
ws=wbout.active; ws.title='文件汇总'
headers=['section','discipline','rel','sheet_count','bill_items','major_amount','major_amount_label','project','sheets','status']
ws.append(headers)
for r in summary_rows:
    ws.append([r.get(h,'') for h in headers])
ws2=wbout.create_sheet('工作表汇总')
headers2=['file','sheet','max_row','max_col','bill_items','amounts']
ws2.append(headers2)
for r in sheet_rows:
    ws2.append([r.get(h,'') for h in headers2])
ws3=wbout.create_sheet('清单项样本')
headers3=['file','sheet','row','code','name','qty_lastnum']
ws3.append(headers3)
for r in bill_rows[:5000]: ws3.append([r.get(h,'') for h in headers3])
# width
for wsx in wbout.worksheets:
    for col in wsx.columns:
        try:
            letter=col[0].column_letter
            wsx.column_dimensions[letter].width=min(max(10, max(len(str(c.value or '')) for c in col[:200])+2), 60)
        except Exception: pass
wbout.save(out_root/'哈密国源工程量清单批量解析.xlsx')

# markdown summary
cnt_section=Counter(r.get('section','') for r in summary_rows)
md=[]
md.append('# 哈密国源原始工程量清单批量解析报告\n\n')
md.append(f'- 解析根路径：`{root}`\n')
md.append(f'- Excel文件数：{len(xlsx_files)}\n')
md.append(f'- 可识别清单项行数：{len(bill_rows)}\n')
md.append(f'- 解析错误：{len(errors)}\n\n')
md.append('## 分类统计\n\n| 分类 | 文件数 |\n|---|---:|\n')
for k,v in cnt_section.most_common(): md.append(f'| {k} | {v} |\n')
md.append('\n## 专业文件汇总\n\n| 序号 | 分类 | 专业 | 文件 | 清单项数 | 主要金额 |\n|---:|---|---|---|---:|---:|\n')
for i,r in enumerate(summary_rows,1):
    amt=r.get('major_amount','')
    amt_s=f"{amt:,.2f}" if isinstance(amt,(int,float)) else ''
    md.append(f"| {i} | {r.get('section','')} | {r.get('discipline','')} | `{r['rel']}` | {r.get('bill_items',0)} | {amt_s} |\n")
if errors:
    md.append('\n## 解析错误\n')
    for e in errors: md.append(f"- `{e['file']}`：{e['error']}\n")
(out_root/'工程量清单批量解析报告.md').write_text(''.join(md),encoding='utf-8')
print('XLSX_FILES', len(xlsx_files))
print('BILL_ROWS', len(bill_rows))
print('ERRORS', len(errors))
print('REPORT', out_root/'哈密国源工程量清单批量解析.xlsx')
